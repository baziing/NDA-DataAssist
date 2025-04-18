import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Protection, Alignment
from openpyxl import Workbook
from .utils import connect_db, connect_db_with_config, execute_query, ensure_dir_exists
from .config import DB_CONFIG
from .config import OUTPUT_DIR, INPUT_DIR
from .config.mail_config import get_mail_config, replace_date_variables
from .email_sender import EmailSender
import os
import time
import logging
from datetime import datetime, timedelta
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from backend.file_name_formatter import format_filename, get_unique_filename
import shutil
import re
import json

UPLOAD_FOLDER = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'tmp'))

def cleanup_old_files(directory, days=7):
    """
    删除指定目录下超过指定天数的文件。
    """
    now = time.time()
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if os.path.isfile(file_path):
            if os.stat(file_path).st_mtime < now - days * 86400:
                try:
                    os.remove(file_path)
                    logging.info(f'删除旧文件: {filename}')
                except Exception as e:
                    logging.error(f'删除文件失败: {filename}, 错误: {e}')

def generate_report(task, task_info, data_frame=None, input_file=None, variables_filename=None, output_file=None, output_dir=None):
    """
    生成报表
    """
    logging.info(f'generate_report called with task: {task}')
    output_path = None  # 初始化 output_path
    try:
        # 初始化setting_info（移到更高的作用域）
        setting_info = {}  # 用于存储setting信息
        
        # 优先使用 data_frame，如果没有，再尝试从 input_file 读取
        if data_frame is None:
            if input_file is None:
                raise ValueError("必须提供 data_frame 或 input_file")
            # 读取输入文件
            input_path = input_file
            # 获取输入文件的所有工作表名称
            xls = pd.ExcelFile(input_path)
            sheet_names = xls.sheet_names
            
            # 读取所有工作表的数据
            all_sheets_data = {}
            
            # 如果是手动上传的Excel文件，从文件中读取setting
            if input_file:
                # 先检查是否存在setting工作表
                setting_sheet = None
                for sheet_name in sheet_names:
                    if '{setting}' in sheet_name.lower():
                        setting_sheet = sheet_name
                        break
                
                # 读取setting工作表
                if setting_sheet:
                    try:
                        setting_df = pd.read_excel(input_path, sheet_name=setting_sheet)
                        # 转换为小写以支持大小写不敏感
                        setting_df.columns = setting_df.columns.str.lower()
                        
                        # 检查必要的列是否存在
                        required_columns = ['fun', 'title', 'config']
                        if all(col in setting_df.columns for col in required_columns):
                            for _, row in setting_df.iterrows():
                                fun = str(row['fun']).strip()
                                title = str(row['title']).strip()
                                config = str(row['config']).strip() if pd.notna(row['config']) else ''
                                
                                if fun == '冻结' and title:
                                    setting_info[title] = config
                    except Exception as e:
                        logging.warning(f'处理setting工作表时出错: {e}')
            
            # 读取其他工作表
            for sheet_name in sheet_names:
                if sheet_name != setting_sheet:  # 跳过setting工作表
                    all_sheets_data[sheet_name] = pd.read_excel(input_path, sheet_name=sheet_name)
            
            # 如果没有工作表，使用默认名称
            if not sheet_names:
                sheet_names = ['汇总报表']
                all_sheets_data = {'汇总报表': pd.DataFrame()}
            
            task.update_progress({'progress': 5, 'log': '读取输入文件'})
        else:
            # 如果使用data_frame，从data_frame中获取sheet信息
            if isinstance(data_frame, dict):
                # 如果data_frame是字典，可能包含多个工作表
                sheet_names = []
                all_sheets_data = {}
                
                # 获取所有工作表信息并按order排序
                sheet_info = []
                for sheet_name, sheet_data in data_frame.items():
                    order = sheet_data.get('order', 999) if isinstance(sheet_data, dict) else 999
                    actual_data = sheet_data.get('data', sheet_data) if isinstance(sheet_data, dict) else sheet_data
                    sheet_info.append((sheet_name, order, actual_data))
                
                # 按order排序
                sheet_info.sort(key=lambda x: x[1])
                
                # 重新组织数据
                for sheet_name, _, sheet_data in sheet_info:
                    sheet_names.append(sheet_name)
                    all_sheets_data[sheet_name] = sheet_data
            else:
                # 如果data_frame是单个DataFrame，使用默认工作表名
                sheet_names = ['汇总报表']
                all_sheets_data = {'汇总报表': data_frame}
            
            # 从task_info中读取setting配置（仅用于定时报表）
            if task and hasattr(task, 'settings'):
                try:
                    logging.info(f'开始处理task的settings')
                    logging.info(f'task类型: {type(task)}')
                    logging.info(f'task属性: {dir(task)}')
                    settings = task.settings
                    logging.info(f'从task读取到的settings: {settings}')
                    if settings:
                        if isinstance(settings, str):
                            logging.info(f'settings是字符串类型，尝试解析')
                            settings = json.loads(settings)
                            logging.info(f'解析后的settings: {settings}')
                        elif isinstance(settings, dict):
                            logging.info(f'settings是字典类型，直接使用')
                            settings = settings
                        
                        # 处理冻结设置
                        if isinstance(settings, dict) and 'freeze' in settings:
                            logging.info(f'发现freeze配置: {settings["freeze"]}')
                            for item in settings['freeze']:
                                if isinstance(item, dict) and 'title' in item and 'config' in item:
                                    title = str(item['title']).strip()
                                    config = str(item['config']).strip() if item['config'] else ''
                                    if title:
                                        setting_info[title] = config
                                        logging.info(f'应用冻结设置: sheet={title}, config={config}')
                except Exception as e:
                    logging.warning(f'处理setting配置时出错: {e}')
                    logging.warning(f'原始settings内容: {task.settings}')
            else:
                logging.info(f'task不存在或没有settings属性')

        # 如果提供了输出目录，使用它；否则使用默认目录
        if output_dir:
            # 确保目录存在
            ensure_dir_exists(output_dir)
        else:
            # 获取当前日期作为目录
            date_str = datetime.now().strftime('%Y%m%d')
            # 创建日期目录
            output_dir = os.path.join('output', 'report-manual', date_str)
            ensure_dir_exists(output_dir)

        # 读取变量文件（如果存在）
        variables = {}
        variables_file = None
        if variables_filename:
            variables_file = os.path.join(UPLOAD_FOLDER, 'variables', variables_filename)
        logging.warning(f'{variables_file}')
        if variables_file and os.path.exists(variables_file):
            try:
                variables_df = pd.read_excel(variables_file)
                # 转换为小写，以支持大小写不敏感的列名
                variables_df.columns = variables_df.columns.str.lower()
                if 'key' in variables_df.columns and 'value' in variables_df.columns:
                    # 使用原始的键名（区分大小写）
                    variables = dict(zip(variables_df['key'], variables_df['value']))
                else:
                    logging.warning('变量文件格式不正确，缺少 key 或 value 列。')
            except Exception as e:
                logging.error(f'读取变量文件失败: {e}')

        # 计算 SQL 查询的总数
        total_queries = sum(len(df) for df in all_sheets_data.values())

        # 创建输出工作簿
        wb = Workbook()
        # 删除默认创建的Sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 为每个工作表创建对应的汇总表并处理SQL查询
        for sheet_name, df in all_sheets_data.items():
            # 创建汇总表
            summary_ws = wb.create_sheet(title=sheet_name)
            
            # 定义默认样式
            default_font = Font(name='微软雅黑', size=12)
            bold_font = Font(name='微软雅黑', size=12, bold=True)
            thin_border = Border(left=Side(style=None), 
                                right=Side(style=None), 
                                top=Side(style=None), 
                                bottom=Side(style=None))
            
            # 处理每一行SQL查询
            summary_row_offset = 1
            pos_dict = {}  # 每个工作表有自己的位置字典
            
            for index, row in df.iterrows():
                if task.cancelled:  # 检查是否取消
                    raise Exception('Task cancelled')
                
                # 获取SQL查询信息
                db_name = row['db_name']
                sql = row['output_sql']
                
                # 检测是否存在sql1, sql2等字段并进行拼接
                sql_index = 1
                sql_parts = [sql]  # 用于记录日志的列表
                while True:
                    sql_field = f'sql{sql_index}'
                    if sql_field in row and row[sql_field] and isinstance(row[sql_field], str):
                        # 添加一个空格作为分隔符，避免SQL语句连接时出现语法错误
                        if not sql.endswith(' ') and not row[sql_field].startswith(' '):
                            sql += ' '
                        sql += row[sql_field]
                        sql_parts.append(row[sql_field])
                        sql_index += 1
                    else:
                        break
                
                # 如果进行了SQL拼接，记录日志
                if len(sql_parts) > 1:
                    total_length = len(sql)
                    logging.info(f'检测到多个SQL片段，已进行拼接: {len(sql_parts)} 个片段，总长度: {total_length} 字符')
                
                format_rules = row.get('format', '')
                transpose = row.get('transpose(Y/N)', '').strip().lower() == 'y'  # 检查是否需要转置
                
                # 创建临时工作表，包含工作表名称
                temp_ws = wb.create_sheet(title=f'临时表_{sheet_name}_{index+1}')
                
                # 从Excel行中获取数据库配置
                db_config = {
                    'host': row.get('db_host', DB_CONFIG['host']),
                    'port': int(row.get('db_port', DB_CONFIG['port'])),
                    'user': row.get('db_user', DB_CONFIG['user']),
                    'password': row.get('db_password', DB_CONFIG['password']),
                    'database': db_name  # 使用指定的数据库名
                }

                # 变量替换
                if variables:
                    for key, value in variables.items():
                        sql = sql.replace('{' + key + '}', str(value))
                
                # 计算当前进度
                sheet_index = list(all_sheets_data.keys()).index(sheet_name)
                completed_queries = sum(len(df) for i, (s, df) in enumerate(all_sheets_data.items()) if i < sheet_index)
                completed_queries += index + 1
                progress = 10 + int((completed_queries / total_queries) * 55)
                task.update_progress({'progress': progress, 'log': f'工作表 {sheet_name} 的第 {index + 1} 个 SQL 开始执行'})
                
                # 连接数据库并执行查询
                connection = connect_db_with_config(db_config)
                try:
                    columns, data = execute_query(connection, sql)
                except Exception as e:
                    # 出错时保存完整SQL到日志
                    error_message = f'SQL执行错误: {str(e)}'
                    task.update_progress({'progress': progress, 'log': error_message})
                    logging.error(error_message)
                    
                    # 将完整SQL保存到单独的日志文件中
                    error_log_dir = os.path.join('logs', 'sql_errors')
                    os.makedirs(error_log_dir, exist_ok=True)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    error_log_file = os.path.join(error_log_dir, f'sql_error_{timestamp}.sql')
                    
                    with open(error_log_file, 'w', encoding='utf-8') as f:
                        f.write(f"-- 错误信息: {str(e)}\n")
                        f.write(f"-- 工作表: {sheet_name}, 行: {index + 1}\n")
                        f.write(f"-- SQL长度: {len(sql)} 字符\n")
                        f.write(f"-- 时间: {timestamp}\n\n")
                        f.write(sql)
                    
                    logging.error(f'完整SQL已保存到: {error_log_file}')
                    raise Exception(f'SQL执行错误: {str(e)}，完整SQL已保存到日志')
                
                if task.cancelled:  # 检查是否取消
                    logging.warning(f'Task cancelled after executing SQL {index + 1}')
                    raise Exception('Task cancelled')

                # 转置处理
                if transpose:
                    try:
                        # 转置
                        data_with_columns = [columns] + data  # 将标题行和数据合并
                        data = list(map(list, zip(*data_with_columns)))  # 转置数据

                        # 删除原始标题行，将第一行数据作为新的标题行
                        columns = data[0]  # 将转置后的第一行作为新的标题行
                        data_without_header = data[1:]  # 移除原始标题行

                        # 重新合并新的标题行和数据
                        data = data_without_header
                        
                    except Exception as e:
                        logging.error(f"转置处理时出错: {e}")
                
                # 写入临时工作表
                temp_row_offset = 1
                # 写入表头
                for col_num, column in enumerate(columns, 1):
                    cell = temp_ws.cell(row=temp_row_offset, column=col_num, value=column)
                    cell.font = bold_font
                    cell.border = thin_border
                    if task.cancelled:  # 检查是否取消
                        raise Exception('Task cancelled')
                
                # 写入数据
                for data_row in data:
                    temp_row_offset += 1
                    for col_num, value in enumerate(data_row, 1):
                        cell = temp_ws.cell(row=temp_row_offset, column=col_num, value=value)
                        cell.font = default_font
                        cell.border = thin_border
                        if task.cancelled:  # 检查是否取消
                            raise Exception('Task cancelled')
                
                # 应用自定义样式
                if format_rules:
                    apply_format_rules(temp_ws, format_rules)
                
                # 更新进度
                task.update_progress({'progress': progress, 'log': f'工作表 {sheet_name} 的第 {index + 1} 个 SQL 应用样式'})
                
                # 解析pos字段
                pos = row.get('pos', '')
                if pos:
                    try:
                        # 解析pos格式 n-m
                        pos_parts = pos.split('-')
                        if len(pos_parts) != 2:
                            raise ValueError('Invalid pos format')
                            
                        pos_row = int(pos_parts[0])
                        pos_col = int(pos_parts[1])
                        
                        if pos_row <= 0 or pos_col <= 0:
                            raise ValueError('Pos values must be positive integers')
                            
                        # 处理位置
                        if pos_row == 1 and pos_col == 1:
                            # 第一个表格，放在左上角
                            start_row = 1
                            start_col = 1
                        else:
                            if pos_col == 1:
                                # 当m_number=1时
                                # 寻找n-1的所有结果的最大行
                                max_row = 0
                                # 遍历所有以n-1开头的pos键
                                for key in pos_dict.keys():
                                    if key.startswith(f'{pos_row-1}-'):
                                        max_row = max(max_row, pos_dict[key]['end_row'])
                                start_row = max_row + 2  # 增加1行间隔
                                start_col = 1
                            else:
                                # 当m_number >= 2时
                                # 找到n=pos_row且m=pos_col-1的结果
                                prev_key = f'{pos_row}-{pos_col-1}'
                                if prev_key in pos_dict:
                                    start_row = pos_dict[prev_key]['start_row']
                                    start_col = pos_dict[prev_key]['end_col'] + 2  # 增加1列间隔
                                else:
                                    # 如果找不到前一个表格，使用默认位置
                                    start_row = summary_row_offset + 1  # 增加1行间隔
                                    start_col = 1
                        
                        # 将当前表格的位置信息存入字典
                        pos_dict[f'{pos_row}-{pos_col}'] = {
                            'start_row': start_row,
                            'start_col': start_col,
                            'end_row': start_row + temp_ws.max_row - 1,
                            'end_col': start_col + temp_ws.max_column - 1
                        }
                        
                        # 增加表格间的空行
                        summary_row_offset += 1
                            
                    except Exception as e:
                        logging.warning(f'Invalid pos value "{pos}": {e}. Using default position.')
                        start_row = summary_row_offset
                        start_col = 1
                else:
                    # 没有pos字段，使用默认顺序
                    start_row = summary_row_offset
                    start_col = 1
                
                # 更新进度
                task.update_progress({'progress': progress, 'log': f'工作表 {sheet_name} 的第 {index + 1} 个 SQL 结果写入汇总表'})
                
                # 将临时表所有数据合并到汇总表，包括格式
                max_row = temp_ws.max_row
                max_col = temp_ws.max_column
                
                for row_idx, row in enumerate(temp_ws.iter_rows(min_row=1, max_row=max_row,
                                           min_col=1, max_col=max_col), start=start_row):
                    for col_idx, cell in enumerate(row, start=start_col):
                        new_cell = summary_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                        # 复制样式
                        if cell.has_style:
                            new_cell.font = Font(
                                name=cell.font.name,
                                size=cell.font.size,
                                bold=cell.font.bold,
                                italic=cell.font.italic,
                                vertAlign=cell.font.vertAlign,
                                underline=cell.font.underline,
                                strike=cell.font.strike,
                                color=cell.font.color
                            )
                            new_cell.border = Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom,
                                diagonal=cell.border.diagonal,
                                diagonal_direction=cell.border.diagonal_direction,
                                outline=cell.border.outline,
                                vertical=cell.border.vertical,
                                horizontal=cell.border.horizontal
                            )
                            new_cell.fill = PatternFill(
                                fill_type=cell.fill.fill_type,
                                start_color=cell.fill.start_color,
                                end_color=cell.fill.end_color
                            )
                            new_cell.number_format = cell.number_format
                            new_cell.protection = Protection(
                                locked=cell.protection.locked,
                                hidden=cell.protection.hidden
                            )
                            new_cell.alignment = Alignment(
                                horizontal=cell.alignment.horizontal,
                                vertical=cell.alignment.vertical,
                                text_rotation=cell.alignment.text_rotation,
                                wrap_text=cell.alignment.wrap_text,
                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                indent=cell.alignment.indent
                            )
                        
                        # 复制条件格式
                        for cf in temp_ws.conditional_formatting:
                            # 获取原始条件格式的范围
                            orig_range = list(cf.cells.ranges)[0]
                            
                            # 计算原始范围的列偏移量
                            orig_start_col = orig_range.min_col
                            orig_end_col = orig_range.max_col
                            
                            # 计算相对范围（保持与原始数据的相对位置关系）
                            relative_start_row = orig_range.min_row  # 保持原始的相对起始行
                            relative_end_row = orig_range.max_row    # 保持原始的相对结束行
                            relative_rows = relative_end_row - relative_start_row  # 计算范围跨度

                            # 计算新范围的起始和结束坐标
                            new_start_row = start_row + (relative_start_row - 1)  # -1 是因为我们要保持相对位置
                            new_end_row = new_start_row + relative_rows
                            new_start_col = start_col + orig_range.min_col - 1
                            new_end_col = start_col + orig_range.max_col - 1
                            
                            # 确保列索引不超过Excel最大列数限制
                            if new_end_col > 16384:  # Excel最大列数为16384
                                new_end_col = 16384
                                new_start_col = new_end_col - (orig_end_col - orig_start_col + 1) + 1
                                if new_start_col < 1:
                                    new_start_col = 1
                            
                            # 创建新的范围字符串，使用get_column_letter函数替代直接使用cell.coordinate
                            from openpyxl.utils import get_column_letter
                            new_start_col_letter = get_column_letter(new_start_col)
                            new_end_col_letter = get_column_letter(new_end_col)
                            new_range = f"{new_start_col_letter}{new_start_row}:{new_end_col_letter}{new_end_row}"
                            
                            # 应用条件格式到指定范围
                            summary_ws.conditional_formatting.add(new_range, cf.rules[0])
                    summary_row_offset += 1
                
                summary_row_offset += 1  # 每个报表之间空一行
            
            # 完成数据写入,开始应用样式之前, 更新进度
            task.update_progress({'progress':70, 'log':f'工作表 {sheet_name} 完成数据写入,开始应用样式'})

            # 应用格式规则到每个临时表
            for index, row in df.iterrows():
                format_rules = row.get('format', '')
                if format_rules:
                    temp_ws_name = f'临时表_{sheet_name}_{index+1}'
                    if temp_ws_name in wb.sheetnames:
                        temp_ws = wb[temp_ws_name]
                        apply_format_rules(temp_ws, format_rules)
                        logging.info(f'应用格式规则到临时表: {temp_ws_name}')

            # 完成数据插入后：新增标题行和首列
            # 保存所有条件格式规则及其范围
            saved_formats = []
            for cf in summary_ws.conditional_formatting:
                for range_obj in cf.cells.ranges:
                    # 保存原始范围的行列信息和相对于数据起始位置的偏移
                    original_start_row = range_obj.min_row - start_row  # 计算相对起始行
                    original_end_row = range_obj.max_row - start_row    # 计算相对结束行
                    original_start_col = range_obj.min_col - start_col  # 计算相对起始列
                    original_end_col = range_obj.max_col - start_col    # 计算相对结束列
                    
                    saved_formats.append((
                        cf.rules[0],
                        original_start_row,  # 保存相对位置
                        original_end_row,
                        original_start_col,
                        original_end_col,
                        start_row,          # 保存原始起始位置
                        start_col
                    ))

            # 清除现有的条件格式
            summary_ws.conditional_formatting = type(summary_ws.conditional_formatting)()

            # 插入新行和列
            summary_ws.insert_rows(1)
            summary_ws.insert_cols(1)

            # 重新应用条件格式，更新范围
            for rule, rel_min_row, rel_max_row, rel_min_col, rel_max_col, orig_start_row, orig_start_col in saved_formats:
                # 计算新的范围坐标，保持相对位置关系
                new_start_row = orig_start_row + rel_min_row + 1  # +1 因为插入了一行
                new_end_row = orig_start_row + rel_max_row + 1
                new_start_col = orig_start_col + rel_min_col + 1  # +1 因为插入了一列
                new_end_col = orig_start_col + rel_max_col + 1
                
                # 使用get_column_letter函数创建新的范围字符串
                from openpyxl.utils import get_column_letter
                start_col_letter = get_column_letter(new_start_col)
                end_col_letter = get_column_letter(new_end_col)
                new_range = f"{start_col_letter}{new_start_row}:{end_col_letter}{new_end_row}"
                
                # 应用更新后的条件格式
                summary_ws.conditional_formatting.add(new_range, rule)
            
            # 完成样式和条件格式写入后, 更新进度
            task.update_progress({'progress':80, 'log':f'工作表 {sheet_name} 完成样式和条件格式写入'})

            # 设置第一行的列宽为3，其他列宽为13
            from openpyxl.utils import get_column_letter
            for col in range(1, summary_ws.max_column + 1):
                col_letter = get_column_letter(col)
                if col == 1:
                    summary_ws.column_dimensions[col_letter].width = 3  # 第一列设置为3
                else:
                    summary_ws.column_dimensions[col_letter].width = 13  # 其他列设置为13
        
        # 完成所有数据插入,新增标题和首列之前, 更新进度
        task.update_progress({'progress':90, 'log':'完成所有数据插入,调整样式'})

        # 根据是否传入 input_file 确定文件名
        if input_file:
            # 手动生成，使用原始文件名
            file_name_without_ext = os.path.splitext(task.original_filename)[0]
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"{file_name_without_ext}_{timestamp}.xlsx"
        else:
            # 定时配置，使用任务名
            # 获取文件名模板，如果没有则使用任务名
            filename_template = getattr(task, 'output_filename', None) or task.task_name
            
            # 尝试格式化文件名
            execution_time = datetime.now()
            formatted_filename, has_replacement = format_filename(filename_template, execution_time)
            
            # 如果没有发生替换，则使用原来的命名规则（任务名+时间戳）
            if not has_replacement:
                timestamp = execution_time.strftime('%Y%m%d_%H%M%S')
                output_file = f"{task.task_name}_{timestamp}.xlsx"
            else:
                # 确保文件名以 .xlsx 结尾
                if not formatted_filename.lower().endswith('.xlsx'):
                    output_file = f"{formatted_filename}.xlsx"
                else:
                    output_file = formatted_filename

        # 确保文件名在输出目录中是唯一的
        output_file = get_unique_filename(output_dir, output_file)

        # 删除所有临时表
        temp_sheets = [sheet for sheet in wb.sheetnames if '临时表_' in sheet]
        for sheet_name in temp_sheets:
            wb.remove(wb[sheet_name])
    
        # 在保存文件之前应用冻结设置
        for sheet_name, freeze_info in setting_info.items():
            try:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if not freeze_info:  # 如果config为空，跳过
                        continue
                        
                    # 处理类似"B2"这样的格式
                    if len(freeze_info) >= 2 and freeze_info[0].isalpha() and freeze_info[1:].isdigit():
                        col_letter = ''.join(c for c in freeze_info if c.isalpha())
                        row_num = int(''.join(c for c in freeze_info if c.isdigit()))
                        ws.freeze_panes = f'{col_letter}{row_num}'
                    # 处理类似"B"这样的格式
                    elif freeze_info.isalpha():
                        ws.freeze_panes = f'{freeze_info.upper()}1'
                    # 处理纯数字格式
                    elif freeze_info.isdigit():
                        row_num = int(freeze_info)
                        ws.freeze_panes = f'A{row_num}'
                    # 处理类似"c2"这样的格式（保持原有逻辑的兼容性）
                    elif freeze_info.startswith('c') and freeze_info[1:].isdigit():
                        col_num = int(freeze_info[1:])
                        ws.freeze_panes = f'{get_column_letter(col_num)}2'
            except Exception as e:
                logging.warning(f'应用冻结时出错: sheet_name={sheet_name}, config={freeze_info}, error={e}')
    
        # 保存文件
        output_path = os.path.join(output_dir, output_file)
        logging.info(f'报表生成路径: {os.path.abspath(output_path)}') # 打印绝对路径
        wb.save(output_path)
        task.update_progress({'progress':100, 'log':'保存文件'}) # 保存文件后：更新 100%
        logging.info(f'报表生成成功: {output_path}')

        # # 获取邮件配置
        # file_name = os.path.basename(input_file)
        # report_type = os.path.basename(os.path.dirname(input_file))
        # logging.info(f'正在获取邮件配置，文件名: {file_name}, 报表类型: {report_type}')
        # mail_config = get_mail_config(file_name, report_type)
        
        # if mail_config:
        #     logging.info(f'找到邮件配置: {mail_config}')
        #     try:
        #         # 替换日期变量
        #         subject = replace_date_variables(mail_config['subject'])
        #         body = replace_date_variables(mail_config['body'])
                
        #         # 发送邮件
        #         sender = EmailSender()
        #         success = sender.send_email(
        #             subject=subject,
        #             recipients=mail_config['to'],
        #             cc=mail_config.get('cc', []),
        #             body=body,
        #             attachments=[output_path]
        #         )
                
        #         if success:
        #             logging.info(f'邮件发送成功: {mail_config["to"]}')
        #         else:
        #             logging.error('邮件发送失败')
        #     except Exception as e:
        #         logging.error(f'邮件发送失败: {e}')
        #     finally:
        #         cleanup_old_files(os.path.join(UPLOAD_FOLDER, 'variables'))
        return output_path
    except Exception as e:
        logging.error(f'报表生成失败: {e}')
        logging.info(f'generate_report returning: {output_path}')
        raise

def apply_format_rules(worksheet, format_rules, max_row_override=None):
    """
    应用自定义样式规则
    """
    try:
        # 确保 format_rules 是字符串
        if not isinstance(format_rules, str):
            format_rules = str(format_rules)
            
        # 解析格式规则
        rules = format_rules.split(';') if format_rules else []
        for rule in rules:
            if not rule.strip():
                continue
                
            # 解析规则类型和参数
            parts = rule.split(':')
            rule_type = parts[0].strip()
            params = parts[1].strip() if len(parts) > 1 else ''
            
            # 处理不同规则类型
            if rule_type == 'bold':
                apply_bold(worksheet, params)
            elif rule_type.startswith('bg_color'):
                apply_bg_color(worksheet, params)
            elif rule_type.startswith('font_color'):
                apply_font_color(worksheet, params)
            elif rule_type.startswith('font'):
                apply_font(worksheet, params)
            elif rule_type == 'border':
                apply_border(worksheet, params)
            elif rule_type.startswith('data_bar'):
                apply_data_bar(worksheet, params)
            elif rule_type.startswith('color_scale'):
                apply_color_scale(worksheet, params)
            elif rule_type.startswith('decimal'):
                apply_decimal_format(worksheet, params)
            elif rule_type == 'alignment':
                apply_alignment(worksheet, params)
            elif rule_type == 'number_format':
                apply_number_format(worksheet, params)
                
    except Exception as e:
        logging.error(f'应用样式规则失败: {e}')
        raise

def apply_alignment(worksheet, params):
    """
    应用对齐样式
    params格式: "对齐方式,起始行-结束行,起始列-结束列"
    对齐方式：left/center/right
    """
    try:
        if not params:
            return
            
        # 解析所有参数
        parts = params.split(',')
        if len(parts) != 3:
            raise ValueError('Invalid alignment format. Expected format: "alignment_type,start_row-end_row,start_col-end_col"')
            
        # 解析对齐方式
        align_type = parts[0].strip().lower()
        if align_type not in ['left', 'center', 'right']:
            raise ValueError('Invalid alignment type. Must be one of: left, center, right')
            
        # 解析行范围
        rows = parts[1].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = max_row_override if max_row_override is not None else worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError(f'Start row ({start_row}) must be less than or equal to end row ({end_row})')
        except ValueError as e:
            raise ValueError(f'Row numbers must be integers or "max". Error: {e}')
            
        # 解析列范围
        cols = parts[2].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建对齐样式
        if align_type == 'left':
            alignment = Alignment(horizontal='left')
        elif align_type == 'center':
            alignment = Alignment(horizontal='center')
        else:
            alignment = Alignment(horizontal='right')
        
        # 应用对齐样式
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = alignment
                
    except Exception as e:
        logging.error(f'应用对齐样式失败: {e}')
        raise

def apply_number_format(worksheet, params):
    """
    应用数字格式
    params格式: "格式类型,小数位数,起始行-结束行,起始列-结束列"
    格式类型：general/number/percentage
    """
    try:
        if not params:
            return
            
        # 解析所有参数
        parts = params.split(',')
        if len(parts) != 4:
            raise ValueError('Invalid number format. Expected format: "format_type,decimal_places,start_row-end_row,start_col-end_col"')
            
        # 解析格式类型
        format_type = parts[0].strip().lower()
        if format_type not in ['general', 'number', 'percentage']:
            raise ValueError('Invalid format type. Must be one of: general, number, percentage')
            
        # 解析小数位数
        try:
            decimal_places = int(parts[1])
            if decimal_places < 0:
                raise ValueError('Decimal places must be a non-negative integer')
        except ValueError:
            raise ValueError('Decimal places must be an integer')
            
        # 解析行范围
        rows = parts[2].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = parts[3].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建数字格式
        if format_type == 'general':
            number_format = 'General'
        elif format_type == 'number':
            number_format = f'0.{decimal_places * "0"}'
        else:
            number_format = f'0.{decimal_places * "0"}%'
        
        # 应用数字格式
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.number_format = number_format
                
    except Exception as e:
        logging.error(f'应用数字格式失败: {e}')
        raise

def apply_font(worksheet, params):
    """
    应用自定义字体样式
    params格式: "font_name,font_size,start_row-end_row,start_col-end_col"
    如果未指定字体名称，则使用默认字体（微软雅黑）
    """
    try:
        if not params:
            return
            
        # 解析所有参数
        parts = params.split(',')
        if len(parts) != 4:
            raise ValueError('Invalid font format. Expected format: "font_name,font_size,start_row-end_row,start_col-end_col"')
            
        # 解析字体名称和大小
        font_name = parts[0].strip() or '微软雅黑'
        font_size = 12
        
        try:
            font_size_str = parts[1].strip()
            if font_size_str:  # 如果提供了字体大小
                font_size = int(font_size_str)
                if font_size <= 0:
                    raise ValueError('Font size must be greater than 0')
        except ValueError:
            raise ValueError('Font size must be a positive integer')
            
        # 解析行范围
        rows = parts[2].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = parts[3].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建字体样式
        font = Font(name=font_name, size=font_size)
        
        # 应用字体样式
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.font:
                    cell.font = cell.font.copy(name=font_name, size=font_size)
                else:
                    cell.font = font
                
    except Exception as e:
        logging.error(f'应用字体样式失败: {e}')
        raise

def apply_bold(worksheet, params):
    """
    应用加粗样式
    params格式: "start_row-end_row,start_col-end_col" 或 "row,col"
    """
    try:
        # 解析行列范围
        if not params:
            return
            
        # 解析区域
        area_parts = params.split(',')
        if len(area_parts) != 2:
            raise ValueError('Invalid bold format')
            
        # 解析行范围
        rows = area_parts[0].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = area_parts[1].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 应用加粗样式
        bold_font = Font(bold=True)
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = bold_font
                
    except Exception as e:
        logging.error(f'应用加粗样式失败: {e}')
        raise

def apply_bg_color(worksheet, params):
    """
    应用背景颜色
    params格式: "color,start_row-end_row,start_col-end_col" 或 "color,row,col"
    """
    try:
        if not params:
            return
            
        # 解析所有参数
        parts = params.split(',')
        if len(parts) != 3:
            raise ValueError('Invalid bg_color format. Expected format: "color,start_row-end_row,start_col-end_col"')
            
        color = parts[0].strip()
        area = f"{parts[1]},{parts[2]}"
        
        # 确保颜色以#开头
        if not color.startswith('#'):
            color = '#' + color
            
        # 确保颜色为6位十六进制
        if len(color) != 7:
            raise ValueError('Invalid color format, must be 6-digit hex')
        
        # 解析行列范围
        area_parts = area.split(',')
        if len(area_parts) != 2:
            raise ValueError('Invalid bg_color area format')
            
        # 解析行范围
        rows = area_parts[0].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = area_parts[1].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建填充样式
        fill = PatternFill(start_color=color[1:],  # 去掉#号
                          end_color=color[1:],
                          fill_type='solid')
        
        # 应用背景颜色
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = fill
                
    except Exception as e:
        logging.warning(f'应用背景颜色失败，使用默认格式: {e}')

def apply_font_color(worksheet, params):
    """
    应用字体颜色
    params格式: "color,start_row-end_row,start_col-end_col" 或 "color,row,col"
    """
    try:
        if not params:
            return
            
        # 解析所有参数
        parts = params.split(',')
        if len(parts) != 3:
            raise ValueError('Invalid font_color format. Expected format: "color,start_row-end_row,start_col-end_col"')
            
        color = parts[0].strip()
        area = f"{parts[1]},{parts[2]}"
        
        # 解析行列范围
        area_parts = area.split(',')
        if len(area_parts) != 2:
            raise ValueError('Invalid font_color area format')
            
        # 解析行范围
        rows = area_parts[0].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = area_parts[1].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建字体样式
        font = Font(color=color[1:])  # 去掉#号
        
        # 应用字体颜色
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.font:
                    cell.font = cell.font.copy(color=color[1:])
                else:
                    cell.font = font
                
    except Exception as e:
        logging.error(f'应用字体颜色失败: {e}')
        raise

def apply_border(worksheet, params):
    """
    应用边框样式
    params格式: "start_row-end_row,start_col-end_col" 或 "row,col"
    """
    try:
        if not params:
            return
            
        # 解析行列范围
        area_parts = params.split(',')
        if len(area_parts) != 2:
            raise ValueError('Invalid border format')
            
        # 解析行范围
        rows = area_parts[0].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = area_parts[1].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建边框样式
        border = Border(left=Side(style='thin'),
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))
        
        # 应用边框样式
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = border
                
    except Exception as e:
        logging.error(f'应用边框样式失败: {e}')
        raise

def apply_data_bar(worksheet, params):
    """
    应用数据条
    params格式: "start_row-end_row,start_col-end_col" 或 "row,col"
    """
    try:
        if not params:
            return
            
        # 解析行列范围
        area_parts = params.split(',')
        if len(area_parts) != 2:
            raise ValueError('Invalid data_bar format')
            
        # 解析行范围
        rows = area_parts[0].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = area_parts[1].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建数据条条件格式
        from openpyxl.formatting.rule import DataBarRule
        rule = DataBarRule(start_type='min',
                          end_type='max',
                          color="638EC6",
                          showValue=True,
                          minLength=None,
                          maxLength=None)
        
        # 使用get_column_letter函数替代直接使用cell.coordinate
        from openpyxl.utils import get_column_letter
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)
        cell_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        
        # 应用数据条
        worksheet.conditional_formatting.add(cell_range, rule)
                
    except Exception as e:
        logging.error(f'应用数据条失败: {e}')
        raise

def apply_color_scale(worksheet, params):
    """
    应用色阶
    params格式: "start_row-end_row,start_col-end_col" 或 "row,col"
    """
    try:
        if not params:
            return
            
        # 解析行列范围
        area_parts = params.split(',')
        if len(area_parts) != 2:
            raise ValueError('Invalid color_scale format')
            
        # 解析行范围
        rows = area_parts[0].split('-')
        if len(rows) != 2:
            raise ValueError('Invalid row range format. Expected format: "start_row-end_row" or "start_row-max"')
            
        try:
            start_row = int(rows[0])
            if start_row <= 0:
                raise ValueError('Start row must be a positive integer')
                
            if rows[1].lower() == 'max':
                end_row = worksheet.max_row
            else:
                end_row = int(rows[1])
                if end_row <= 0:
                    raise ValueError('End row must be a positive integer')
                if start_row > end_row:
                    raise ValueError('Start row must be less than or equal to end row')
        except ValueError:
            raise ValueError('Row numbers must be integers or "max"')
            
        # 解析列范围
        cols = area_parts[1].split('-')
        if len(cols) != 2:
            raise ValueError('Invalid column range format. Expected format: "start_col-end_col" or "start_col-max"')
            
        try:
            start_col = int(cols[0])
            if start_col <= 0:
                raise ValueError('Start column must be a positive integer')
                
            if cols[1].lower() == 'max':
                end_col = worksheet.max_column
            else:
                end_col = int(cols[1])
                if end_col <= 0:
                    raise ValueError('End column must be a positive integer')
                if start_col > end_col:
                    raise ValueError('Start column must be less than or equal to end column')
        except ValueError:
            raise ValueError('Column numbers must be integers or "max"')
        
        # 创建色阶条件格式
        from openpyxl.formatting.rule import ColorScaleRule
        rule = ColorScaleRule(start_type='min',
                             start_color='FCFCFF',
                            #  mid_type='percentile',
                            #  mid_color='FFEB84',  # 中间颜色暂不支持，可以自己修改
                             end_type='max',
                             end_color='63BE7B')
        
        # 使用get_column_letter函数替代直接使用cell.coordinate
        from openpyxl.utils import get_column_letter
        start_cell_col = get_column_letter(start_col)
        end_cell_col = get_column_letter(end_col)
        cell_range = f"{start_cell_col}{start_row}:{end_cell_col}{end_row}"
        
        # 应用色阶到指定范围
        logging.info(f"Applying color scale with range: {cell_range}, start_row: {start_row}, end_row: {end_row}, start_col: {start_col}, end_col: {end_col}")  # Add logging
        worksheet.conditional_formatting.add(cell_range, rule)

    except Exception as e:
        logging.error(f'应用色阶失败: {e}')
        raise
