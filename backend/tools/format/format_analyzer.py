import openpyxl
import logging
import sys
import os
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

class ExcelFormatAnalyzer:
    def __init__(self):
        # 配置日志
        logging.basicConfig(
            level=logging.INFO,
            format='%(message)s',
            stream=sys.stdout
        )
        
        # 设置默认输出目录
        self.default_output_dir = os.path.join('tools', 'format', 'output')
        
    def analyze_excel(self, excel_file, sheet_name=None):
        """
        分析Excel文件并生成格式规则
        """
        try:
            logging.info(f"正在分析文件: {excel_file}")
            
            # 获取文件名（不含路径）
            file_name = os.path.basename(excel_file)
            
            # 加载工作簿
            wb = openpyxl.load_workbook(excel_file)
            
            # 如果没有指定sheet_name，则分析所有工作表
            sheets_to_analyze = [wb[sheet_name]] if sheet_name else wb.worksheets
            
            # 存储所有工作表的规则
            all_sheet_rules = []
            
            for ws in sheets_to_analyze:
                logging.info(f"正在处理工作表: {ws.title}")
                
                # 初始化样式容器
                style_positions = {
                    'font': {},
                    'bold': set(),
                    'bg_color': {},
                    'font_color': {},
                    'border': set(),
                    'alignment': {},
                    'number_format': {}
                }
                
                # 分析每个单元格的样式
                for row in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        self._analyze_cell_style(cell, row, col, style_positions)
                
                # 分析条件格式
                conditional_formats = self._analyze_conditional_formatting(ws)
                
                # 生成规则
                rules = self._generate_rules(style_positions, conditional_formats)
                
                # 合并当前工作表的所有规则
                sheet_rules = ';'.join(filter(None, rules))
                
                # 如果有规则，添加到结果列表，使用新的格式：【文件名 工作表名】
                if sheet_rules:
                    all_sheet_rules.append(f"【{file_name} {ws.title}】{sheet_rules}")
            
            # 用换行符连接所有工作表的规则
            final_rules = '\n'.join(all_sheet_rules)
            
            logging.info("分析完成，生成的规则：")
            logging.info(final_rules)
            
            return final_rules
            
        except Exception as e:
            logging.error(f'分析Excel样式失败: {e}')
            raise

    def _analyze_cell_style(self, cell, row, col, style_positions):
        """分析单个单元格的样式"""
        # 分析字体
        if cell.font:
            # 字体名称和大小
            if cell.font.name and cell.font.size:  # 添加空值检查
                font_key = f"{cell.font.name},{cell.font.size}"
                if font_key not in style_positions['font']:
                    style_positions['font'][font_key] = set()
                style_positions['font'][font_key].add((row, col))
            
            # 加粗
            if cell.font.bold:
                style_positions['bold'].add((row, col))
            
            # 字体颜色
            if cell.font.color and cell.font.color.rgb:
                color = cell.font.color.rgb
                if color != '00000000':  # 忽略默认颜色
                    if color not in style_positions['font_color']:
                        style_positions['font_color'][color] = set()
                    style_positions['font_color'][color].add((row, col))
        
        # 分析背景色
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            color = cell.fill.start_color.rgb
            if color != '00000000':  # 忽略默认填充
                if color not in style_positions['bg_color']:
                    style_positions['bg_color'][color] = set()
                style_positions['bg_color'][color].add((row, col))
        
        # 分析边框
        if cell.border and any([cell.border.left.style, cell.border.right.style,
                              cell.border.top.style, cell.border.bottom.style]):
            style_positions['border'].add((row, col))
        
        # 分析对齐方式
        if cell.alignment and cell.alignment.horizontal:
            align_type = cell.alignment.horizontal
            if align_type not in style_positions['alignment']:
                style_positions['alignment'][align_type] = set()
            style_positions['alignment'][align_type].add((row, col))
        
        # 分析数字格式
        if cell.number_format != 'General':
            if cell.number_format not in style_positions['number_format']:
                style_positions['number_format'][cell.number_format] = set()
            style_positions['number_format'][cell.number_format].add((row, col))

    def _analyze_conditional_formatting(self, worksheet):
        """分析条件格式"""
        conditional_formats = []
        for cf in worksheet.conditional_formatting:
            rule = cf.rules[0]
            if isinstance(rule, DataBarRule):
                ranges = [str(r) for r in cf.cells.ranges]
                conditional_formats.append(f"data_bar:{ranges[0]}")
            elif isinstance(rule, ColorScaleRule):
                ranges = [str(r) for r in cf.cells.ranges]
                conditional_formats.append(f"color_scale:{ranges[0]}")
        return conditional_formats

    def _generate_rules(self, style_positions, conditional_formats):
        """生成格式规则"""
        rules = []
        
        # 生成字体规则
        for font_key, positions in style_positions['font'].items():
            font_name, font_size = font_key.split(',')
            range_str = self._convert_positions_to_range(positions)
            if range_str:
                rules.append(f"font:{font_name},{font_size},{range_str}")
        
        # 生成加粗规则
        bold_range = self._convert_positions_to_range(style_positions['bold'])
        if bold_range:
            rules.append(f"bold:{bold_range}")
        
        # 生成背景颜色规则
        for color, positions in style_positions['bg_color'].items():
            range_str = self._convert_positions_to_range(positions)
            if range_str:
                rules.append(f"bg_color:#{color},{range_str}")
        
        # 生成字体颜色规则
        for color, positions in style_positions['font_color'].items():
            range_str = self._convert_positions_to_range(positions)
            if range_str:
                rules.append(f"font_color:#{color},{range_str}")
        
        # 生成边框规则
        border_range = self._convert_positions_to_range(style_positions['border'])
        if border_range:
            rules.append(f"border:{border_range}")
        
        # 生成对齐规则
        for align_type, positions in style_positions['alignment'].items():
            range_str = self._convert_positions_to_range(positions)
            if range_str:
                rules.append(f"alignment:{align_type},{range_str}")
        
        # 生成数字格式规则
        for format_str, positions in style_positions['number_format'].items():
            range_str = self._convert_positions_to_range(positions)
            if range_str:
                if '%' in format_str:
                    decimal_places = len(format_str.split('.')[1].replace('%', '')) if '.' in format_str else 0
                    rules.append(f"number_format:percentage,{decimal_places},{range_str}")
                elif '0' in format_str:
                    decimal_places = len(format_str.split('.')[1]) if '.' in format_str else 0
                    rules.append(f"number_format:number,{decimal_places},{range_str}")
        
        # 添加条件格式规则
        rules.extend(conditional_formats)
        
        return rules

    def _convert_positions_to_range(self, positions):
        """将单元格位置集合转换为范围字符串"""
        if not positions:
            return ''
        
        rows = [pos[0] for pos in positions]
        cols = [pos[1] for pos in positions]
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)
        
        return f"{min_row}-{max_row},{min_col}-{max_col}"

    def ensure_output_dir(self):
        """确保输出目录存在"""
        os.makedirs(self.default_output_dir, exist_ok=True)

    def get_default_output_path(self):
        """获取默认输出文件路径"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return os.path.join(self.default_output_dir, f'format_rules_{timestamp}.txt')

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='从Excel文件生成格式规则')
    parser.add_argument('input', help='输入Excel文件路径')
    parser.add_argument('-s', '--sheet', help='工作表名称（可选）')
    parser.add_argument('-o', '--output', help='输出文件路径（可选）')
    
    args = parser.parse_args()
    
    try:
        analyzer = ExcelFormatAnalyzer()
        
        # 分析Excel文件
        format_rules = analyzer.analyze_excel(args.input, args.sheet)
        
        # 确定输出路径
        if args.output:
            output_path = args.output
        else:
            # 确保输出目录存在
            analyzer.ensure_output_dir()
            # 使用默认输出路径
            output_path = analyzer.get_default_output_path()
        
        # 确保输出文件的目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # 写入结果
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(format_rules)
        logging.info(f"结果已保存到: {output_path}")
            
    except Exception as e:
        sys.stderr.write(f"错误: {e}\n")
        sys.stderr.flush()
        sys.exit(1)

if __name__ == "__main__":
    main()
